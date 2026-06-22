#!/usr/bin/env python3
"""
圆柱锂电池全参数设计程序 v4.0
5份行业资料 · 阿基米德螺线 · SOC100%膨胀 · 电解液 & 辅件质量
"""

import sys, os, math, threading
ROOT = os.path.dirname(os.path.abspath(__file__)); sys.path.insert(0, ROOT)

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext

from core.electrode_calc import DesignInput, FullCellResult, run_full_cell_design
from core.material_db import (POSITIVE_MATERIALS, NEGATIVE_MATERIALS, CELL_SPECS,
                               blend_material, ADDITIVE_DENSITY)

IOS = {"bg":"#F2F2F7","card":"#FFFFFF","fg":"#1C1C1E","secondary":"#8E8E93",
       "accent":"#007AFF","accent2":"#5856D6","success":"#34C759","warning":"#FF9500",
       "danger":"#FF3B30","border":"#E5E5EA","input_bg":"#F9F9F9"}
F_T,F_H,F_N,F_S=("Microsoft YaHei UI",16,"bold"),("Microsoft YaHei UI",11,"bold"),("Microsoft YaHei UI",10),("Microsoft YaHei UI",9)
F_MONO=("Consolas",9)

DEFAULTS = {"target_cap":"3000","pos_rev_cap":"182","pos_charge_cap":"191.6","pos_eff":"0.95",
    "pos_true_d":"4.65","pos_act_f":"0.95","pos_sp_f":"0.02","pos_cnt_f":"0.0","pos_bind_f":"0.03",
    "pos_ad":"20","pos_pd":"3.45","pos_foil":"15","pos_width":"57.0","pos_exp":"0.0",
    "neg_rev_cap":"350","neg_charge_cap":"380.4","neg_eff":"0.92","neg_true_d":"2.2",
    "neg_act_f":"0.95","neg_sp_f":"0.01","neg_cnt_f":"0.0","neg_cmc_f":"0.015","neg_bind_f":"0.025",
    "neg_ad":"11","neg_pd":"1.65","neg_foil":"8","neg_wextra":"2.0","neg_exp":"0.01",
    "np_target":"1.10","sep_t":"16","sep_por":"0.49","sep_wextra":"2.0","sep_n1":"2","sep_n2":"2",
    "mandrel":"3.2","neg_lextra":"30","clearance":"0.4","elyte_d":"1.22","elyte_f":"1.20",
    "al_elong":"0.005","cu_elong":"0.0","sback":"0.01"}


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("圆柱锂电池全参数设计 v4.0 — 18650·21700·4680·33140")
        self.root.geometry("1300x900"); self.root.minsize(1100,750)
        self.root.state('zoomed'); self.root.configure(bg=IOS["bg"])
        self._setup_style(); self.result = None; self._build_ui()

    def _setup_style(self):
        s=ttk.Style();s.theme_use("clam")
        s.configure("TNotebook",background=IOS["bg"],borderwidth=0)
        s.configure("TNotebook.Tab",background=IOS["card"],foreground=IOS["fg"],font=F_N,padding=[14,4])
        s.map("TNotebook.Tab",background=[("selected",IOS["card"])],foreground=[("selected",IOS["accent"])])
        s.configure("TFrame",background=IOS["bg"])
        s.configure("TLabel",background=IOS["card"],foreground=IOS["fg"],font=F_N)
        s.configure("TEntry",fieldbackground=IOS["input_bg"],foreground=IOS["fg"],font=F_N)
        s.configure("TCombobox",fieldbackground=IOS["input_bg"],foreground=IOS["fg"],font=F_N)
        s.configure("TButton",background=IOS["card"],foreground=IOS["accent"],font=F_N,borderwidth=1,padding=(8,3))
        s.map("TButton",background=[("active","#E8F0FE")])
        s.configure("Primary.TButton",background=IOS["accent"],foreground="white",font=("Microsoft YaHei UI",9,"bold"),borderwidth=0,padding=(14,6))
        s.map("Primary.TButton",background=[("active","#0056CC")])
        s.configure("Danger.TButton",background=IOS["danger"],foreground="white",font=F_N,borderwidth=0,padding=(8,3))

    def _card(self,parent): return tk.Frame(parent,bg=IOS["card"],highlightthickness=0)

    def _build_ui(self):
        h=tk.Frame(self.root,bg=IOS["card"],height=42);h.pack(fill=tk.X);h.pack_propagate(False)
        tk.Label(h,text="圆柱锂电池全参数设计 v4.0",font=F_T,fg=IOS["fg"],bg=IOS["card"]).pack(side=tk.LEFT,padx=18,pady=4)
        tk.Label(h,text="18650 · 21700 · 4680 · 33140  |  全物理参数 · SOC100%膨胀 · 辅件质量",font=F_S,fg=IOS["secondary"],bg=IOS["card"]).pack(side=tk.LEFT,padx=(0,18),pady=10)
        nb=ttk.Notebook(self.root);nb.pack(fill=tk.BOTH,expand=True,padx=6,pady=4)
        self.root.rowconfigure(1,weight=1);self.root.columnconfigure(0,weight=1)
        self._build_input_tab(nb)
        self._build_result_tab(nb)
        self._build_detail_tab(nb)
        st=tk.Frame(self.root,bg=IOS["card"],height=18);st.pack(fill=tk.X);st.pack_propagate(False)
        self.status=tk.Label(st,text="  就绪",font=F_S,fg=IOS["secondary"],bg=IOS["card"],anchor=tk.W)
        self.status.pack(side=tk.LEFT,fill=tk.X,padx=10)

    # ═══ Tab 1: 输入 ═══
    def _build_input_tab(self,nb):
        tab=ttk.Frame(nb);nb.add(tab,text="  设计计算  ")
        tab.columnconfigure(0,weight=2);tab.columnconfigure(1,weight=3);tab.rowconfigure(0,weight=1)

        # === LEFT: scrollable input ===
        left=tk.Frame(tab,bg=IOS["bg"]);left.grid(row=0,column=0,sticky="nsew",padx=(4,2),pady=4)
        cv=tk.Canvas(left,bg=IOS["bg"],highlightthickness=0)
        sb=ttk.Scrollbar(left,orient=tk.VERTICAL,command=cv.yview)
        sf=tk.Frame(cv,bg=IOS["bg"])
        sf.bind("<Configure>",lambda e:cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0,0),window=sf,anchor=tk.NW)
        cv.configure(yscrollcommand=sb.set);cv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True);sb.pack(side=tk.RIGHT,fill=tk.Y)

        self.iv={};rr=[0]
        def sec(t):
            tk.Label(sf,text=t,font=F_H,fg=IOS["accent"],bg=IOS["bg"]).grid(row=rr[0],column=0,columnspan=4,sticky=tk.W,pady=(6,1),padx=6);rr[0]+=1
        def fld(label,vname,default,unit="",c=0,w=9):
            tk.Label(sf,text=label,font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).grid(row=rr[0],column=c*2,sticky=tk.W,padx=(6,1),pady=0)
            var=tk.StringVar(value=str(default));self.iv[vname]=var
            f=tk.Frame(sf,bg=IOS["bg"]);f.grid(row=rr[0],column=c*2+1,sticky=tk.EW,padx=(1,4),pady=0)
            tk.Entry(f,textvariable=var,bg=IOS["input_bg"],fg=IOS["fg"],font=F_N,width=w,relief=tk.SOLID,borderwidth=1).pack(side=tk.LEFT)
            if unit:tk.Label(f,text=unit,font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT,padx=(1,0))
        def frow(l1,v1,d1,u1,l2,v2,d2,u2):fld(l1,v1,d1,u1,0);fld(l2,v2,d2,u2,1);rr[0]+=1

        sec("电芯规格 & 材料")
        fp=tk.Frame(sf,bg=IOS["bg"]);fp.grid(row=rr[0],column=0,columnspan=4,sticky=tk.EW,padx=6,pady=1);rr[0]+=1
        tk.Label(fp,text="规格",font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT)
        self.cell_type=tk.StringVar(value="18650")
        ttk.Combobox(fp,textvariable=self.cell_type,values=["18650","21700","33140","4680"],state="readonly",width=7).pack(side=tk.LEFT,padx=3)
        tk.Label(fp,text="正极",font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT,padx=(6,0))
        self.pos_mat=tk.StringVar(value="NCM811")
        pos_cb=ttk.Combobox(fp,textvariable=self.pos_mat,values=list(POSITIVE_MATERIALS.keys()),state="readonly",width=7)
        pos_cb.pack(side=tk.LEFT,padx=2);pos_cb.bind("<<ComboboxSelected>>",lambda e:self._load_mat())
        tk.Label(fp,text="负极",font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT,padx=(4,0))
        self.neg_mat=tk.StringVar(value="天然石墨")
        neg_cb=ttk.Combobox(fp,textvariable=self.neg_mat,values=list(NEGATIVE_MATERIALS.keys()),state="readonly",width=10)
        neg_cb.pack(side=tk.LEFT,padx=2);neg_cb.bind("<<ComboboxSelected>>",lambda e:self._load_mat())
        ttk.Button(fp,text="加载",command=self._load_mat).pack(side=tk.LEFT,padx=(4,0))

        fb=tk.Frame(sf,bg=IOS["bg"]);fb.grid(row=rr[0],column=0,columnspan=4,sticky=tk.EW,padx=6,pady=1);rr[0]+=1
        tk.Label(fb,text="正极掺杂",font=F_S,fg=IOS["warning"],bg=IOS["bg"]).pack(side=tk.LEFT)
        self.pbe=tk.BooleanVar();tk.Checkbutton(fb,variable=self.pbe,font=F_S,bg=IOS["bg"],fg=IOS["accent"],selectcolor=IOS["bg"],activebackground=IOS["bg"],command=self._load_mat).pack(side=tk.LEFT)
        self.pbm=tk.StringVar(value="NCM811");ttk.Combobox(fb,textvariable=self.pbm,values=list(POSITIVE_MATERIALS.keys()),state="readonly",width=7).pack(side=tk.LEFT,padx=2)
        self.pbr=tk.StringVar(value="80");tk.Entry(fb,textvariable=self.pbr,bg=IOS["input_bg"],fg=IOS["fg"],font=F_N,width=3,relief=tk.SOLID,borderwidth=1).pack(side=tk.LEFT)
        tk.Label(fb,text="%",font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT)
        tk.Label(fb,text=" 负极掺杂",font=F_S,fg=IOS["warning"],bg=IOS["bg"]).pack(side=tk.LEFT,padx=(4,0))
        self.nbe=tk.BooleanVar();tk.Checkbutton(fb,variable=self.nbe,font=F_S,bg=IOS["bg"],fg=IOS["accent"],selectcolor=IOS["bg"],activebackground=IOS["bg"],command=self._load_mat).pack(side=tk.LEFT)
        self.nbm=tk.StringVar(value="石墨+Si (96.5/3.5)");ttk.Combobox(fb,textvariable=self.nbm,values=list(NEGATIVE_MATERIALS.keys()),state="readonly",width=12).pack(side=tk.LEFT,padx=2)
        self.nbr=tk.StringVar(value="95");tk.Entry(fb,textvariable=self.nbr,bg=IOS["input_bg"],fg=IOS["fg"],font=F_N,width=3,relief=tk.SOLID,borderwidth=1).pack(side=tk.LEFT)
        tk.Label(fb,text="%",font=F_S,fg=IOS["secondary"],bg=IOS["bg"]).pack(side=tk.LEFT)

        sec("正极 (NCM811)")
        frow("目标容量","target_cap","3000","mAh","克容量","pos_rev_cap","160","mAh/g")
        frow("充电容量","pos_charge_cap","168.4","mAh/g","首效","pos_eff","0.95","")
        frow("真密度","pos_true_d","4.65","g/cc","活物质%","pos_act_f","0.94","")
        frow("SP%","pos_sp_f","0.02","","CNT%","pos_cnt_f","0.0","")
        frow("PVdF%","pos_bind_f","0.03","","面密度","pos_ad","20","mg/cm²")
        frow("压实密度","pos_pd","3.5","g/cc","Al箔厚","pos_foil","15","µm")
        frow("敷料宽度","pos_width","57.0","mm","膨胀率","pos_exp","0.0","")

        sec("负极 (石墨)")
        frow("克容量","neg_rev_cap","350","mAh/g","充电容量","neg_charge_cap","380.4","mAh/g")
        frow("首效","neg_eff","0.92","","","","","")
        frow("真密度","neg_true_d","2.2","g/cc","活物质%","neg_act_f","0.95","")
        frow("SP%","neg_sp_f","0.01","","CNT%","neg_cnt_f","0.0","")
        frow("CMC%","neg_cmc_f","0.015","","SBR%","neg_bind_f","0.025","")
        frow("面密度","neg_ad","11","mg/cm²","压实密度","neg_pd","1.65","g/cc")
        frow("Cu箔厚","neg_foil","8","µm","额外宽","neg_wextra","2.0","mm")
        frow("长度余量","neg_lextra","30","mm","膨胀率","neg_exp","0.01","")

        sec("隔膜·卷绕·电解液")
        frow("隔膜厚","sep_t","16","µm","孔隙率","sep_por","0.49","")
        frow("隔膜外宽","sep_wextra","2.0","mm","先卷n1","sep_n1","2","圈")
        frow("后卷n2","sep_n2","2","圈","卷针Ø","mandrel","3.2","mm")
        frow("N/P目标","np_target","1.10","","壳间隙","clearance","0.4","mm")
        frow("电解液ρ","elyte_d","1.22","g/cc","注入系数","elyte_f","1.20","×")
        frow("Al延展","al_elong","0.005","","Cu延展","cu_elong","0.0","")
        frow("回弹率","sback","0.01","","","","","")

        bf=tk.Frame(sf,bg=IOS["bg"]);bf.grid(row=rr[0],column=0,columnspan=4,pady=(8,4),padx=6);rr[0]+=1
        ttk.Button(bf,text="⚡ 开始计算",style="Primary.TButton",command=self._calc).pack(side=tk.LEFT,padx=(0,3))
        ttk.Button(bf,text="清空",style="Danger.TButton",command=self._clear).pack(side=tk.LEFT,padx=2)
        ttk.Button(bf,text="导出Excel",command=self._export).pack(side=tk.LEFT,padx=2)

        # === RIGHT: table output ===
        right=tk.Frame(tab,bg=IOS["bg"]);right.grid(row=0,column=1,sticky="nsew",padx=(2,4),pady=4)
        right.rowconfigure(0,weight=1);right.columnconfigure(0,weight=1)
        c=self._card(right);c.grid(row=0,column=0,sticky="nsew")
        c.rowconfigure(0,weight=0);c.rowconfigure(1,weight=1);c.columnconfigure(0,weight=1)
        tk.Label(c,text="  设计计算结果",font=F_H,fg=IOS["accent"],bg=IOS["card"]).grid(row=0,column=0,sticky=tk.W,padx=10,pady=(6,2))

        # Treeview table (compact, fit one screen)
        cols=("参数","数值","单位")
        self.rpt_tree=ttk.Treeview(c,columns=cols,show="headings",height=20)
        s_=ttk.Style();s_.configure("Treeview",font=("Microsoft YaHei UI",9),rowheight=22)
        s_.configure("Treeview.Heading",font=("Microsoft YaHei UI",9,"bold"))
        for col in cols:self.rpt_tree.heading(col,text=col);self.rpt_tree.column(col,width=110,anchor="center")
        self.rpt_tree.column("参数",width=180,anchor="w")
        self.rpt_tree.grid(row=1,column=0,sticky="nsew",padx=8,pady=(2,8))
        # scrollbar
        tvsb=ttk.Scrollbar(c,orient=tk.VERTICAL,command=self.rpt_tree.yview)
        tvsb.grid(row=1,column=1,sticky="ns",pady=(2,8))
        self.rpt_tree.configure(yscrollcommand=tvsb.set)

    # ═══ Tab 2: 结果 ═══
    def _build_result_tab(self,nb):
        tab=ttk.Frame(nb);nb.add(tab,text="  计算结果  ")
        tab.columnconfigure(0,weight=1);tab.columnconfigure(1,weight=1)
        tab.rowconfigure(0,weight=1);tab.rowconfigure(1,weight=1)
        self.panes={}
        for i,(t,k) in enumerate([("正极参数","pos"),("负极参数","neg"),("卷绕 & 隔膜","winding"),("电解液 & 汇总","summary")]):
            c=self._card(tab);c.grid(row=i//2,column=i%2,sticky="nsew",padx=3,pady=3)
            c.rowconfigure(1,weight=1);c.columnconfigure(0,weight=1)
            tk.Label(c,text=f"  {t}",font=F_H,fg=IOS["accent"],bg=IOS["card"]).pack(anchor=tk.W,padx=10,pady=(4,0))
            txt=scrolledtext.ScrolledText(c,font=F_MONO,bg="#FAFAFA",fg=IOS["fg"],relief=tk.FLAT,wrap=tk.WORD)
            txt.pack(fill=tk.BOTH,expand=True,padx=6,pady=(2,6))
            self.panes[k]=txt

    # ═══ Tab 3: 详情 ═══
    def _build_detail_tab(self,nb):
        tab=ttk.Frame(nb);nb.add(tab,text="  配方详情  ")
        tab.columnconfigure(0,weight=1);tab.columnconfigure(1,weight=1)
        tab.rowconfigure(0,weight=1)
        for i,(t,k) in enumerate([("正极配方 & 质量","pos_detail"),("负极配方 & 质量","neg_detail")]):
            c=self._card(tab);c.grid(row=0,column=i,sticky="nsew",padx=3,pady=3)
            c.rowconfigure(1,weight=1);c.columnconfigure(0,weight=1)
            tk.Label(c,text=f"  {t}",font=F_H,fg=IOS["accent"],bg=IOS["card"]).pack(anchor=tk.W,padx=10,pady=(4,0))
            txt=scrolledtext.ScrolledText(c,font=F_MONO,bg="#FAFAFA",fg=IOS["fg"],relief=tk.FLAT,wrap=tk.WORD)
            txt.pack(fill=tk.BOTH,expand=True,padx=6,pady=(2,6))
            self.panes[k]=txt

    # ═══ Logic ═══
    def _read_input(self):
        v=self.iv;inp=DesignInput()
        def g(k,d): return v[k].get() if k in v else str(d)
        inp.cell_type=self.cell_type.get()
        inp.target_capacity_mAh=float(g("target_cap",3000))
        inp.pos_reversible_capacity_mAh_g=float(g("pos_rev_cap",160))
        inp.pos_charge_capacity_mAh_g=float(g("pos_charge_cap",168.4))
        inp.pos_first_efficiency=float(g("pos_eff",0.95))
        inp.pos_true_density_g_cm3=float(g("pos_true_d",4.65))
        inp.pos_active_fraction=float(g("pos_act_f",0.94))
        inp.pos_SP_fraction=float(g("pos_sp_f",0.02))
        inp.pos_CNT_fraction=float(g("pos_cnt_f",0.0))
        inp.pos_binder_fraction=float(g("pos_bind_f",0.03))
        inp.pos_coating_ad_mg_cm2=float(g("pos_ad",20))
        inp.pos_pressed_density_g_cm3=float(g("pos_pd",3.5))
        inp.pos_foil_thickness_um=float(g("pos_foil",15))
        inp.pos_width_mm=float(g("pos_width",57.0))
        inp.pos_coating_expansion=float(g("pos_exp",0.0))
        inp.neg_reversible_capacity_mAh_g=float(g("neg_rev_cap",350))
        inp.neg_charge_capacity_mAh_g=float(g("neg_charge_cap",380.4))
        inp.neg_first_efficiency=float(g("neg_eff",0.92))
        inp.neg_true_density_g_cm3=float(g("neg_true_d",2.2))
        inp.neg_active_fraction=float(g("neg_act_f",0.95))
        inp.neg_SP_fraction=float(g("neg_sp_f",0.01))
        inp.neg_CNT_fraction=float(g("neg_cnt_f",0.0))
        inp.neg_CMC_fraction=float(g("neg_cmc_f",0.015))
        inp.neg_binder_fraction=float(g("neg_bind_f",0.025))
        inp.neg_coating_ad_mg_cm2=float(g("neg_ad",11))
        inp.neg_pressed_density_g_cm3=float(g("neg_pd",1.65))
        inp.neg_foil_thickness_um=float(g("neg_foil",8))
        inp.neg_width_extra_mm=float(g("neg_wextra",2.0))
        inp.neg_coating_expansion=float(g("neg_exp",0.01))
        inp.np_ratio_target=float(g("np_target",1.10))
        inp.separator_thickness_um=float(g("sep_t",16))
        inp.separator_porosity=float(g("sep_por",0.49))
        inp.separator_width_extra_mm=float(g("sep_wextra",2.0))
        inp.separator_pre_wind_n1=int(float(g("sep_n1",2)))
        inp.separator_post_wind_n2=int(float(g("sep_n2",2)))
        inp.mandrel_diameter_mm=float(g("mandrel",3.2))
        inp.neg_length_extra_mm=float(g("neg_lextra",30))
        inp.core_clearance_mm=float(g("clearance",0.4))
        inp.electrolyte_density_g_cm3=float(g("elyte_d",1.22))
        inp.electrolyte_injection_factor=float(g("elyte_f",1.37))
        inp.al_foil_elongation=float(g("al_elong",0.005))
        inp.cu_foil_elongation=float(g("cu_elong",0.0))
        inp.coating_springback=float(g("sback",0.01))
        return inp

    def _load_mat(self):
        pos=POSITIVE_MATERIALS.get(self.pos_mat.get());neg=NEGATIVE_MATERIALS.get(self.neg_mat.get())
        if self.pbe.get():
            try:r=max(0,min(1,float(self.pbr.get())/100))
            except:r=0.8
            pos=blend_material(pos,POSITIVE_MATERIALS[self.pbm.get()],r)
        if self.nbe.get():
            try:r=max(0,min(1,float(self.nbr.get())/100))
            except:r=0.95
            neg=blend_material(neg,NEGATIVE_MATERIALS[self.nbm.get()],r)
        if pos:
            for k,v in [("pos_rev_cap","reversible_capacity_mAh_g"),("pos_charge_cap","charge_capacity_mAh_g"),
                         ("pos_eff","first_efficiency"),("pos_true_d","true_density_g_cm3"),
                         ("pos_act_f","active_fraction"),("pos_pd","pressed_density_g_cm3")]:
                if v in pos:self.iv[k].set(str(pos[v]))
        if neg:
            for k,v in [("neg_rev_cap","reversible_capacity_mAh_g"),("neg_charge_cap","charge_capacity_mAh_g"),
                         ("neg_eff","first_efficiency"),("neg_true_d","true_density_g_cm3"),
                         ("neg_act_f","active_fraction"),("neg_pd","pressed_density_g_cm3")]:
                if v in neg:self.iv[k].set(str(neg[v]))
        spec=CELL_SPECS[self.cell_type.get()];self.iv["mandrel"].set(str(sum(spec["mandrel_range"])/2))

    def _clear(self):
        for k,v in DEFAULTS.items():
            if k in self.iv:self.iv[k].set(v)
        try:self.rpt_tree.delete(*self.rpt_tree.get_children())
        except:pass
        for p in self.panes.values():p.delete(1.0,tk.END)

    def _calc(self):
        try:inp=self._read_input()
        except ValueError as e:messagebox.showerror("输入错误",str(e));return
        try:
            r=run_full_cell_design(inp);self.result=r
            self._show_summary(inp,r)
            try:self._show_panes(r)
            except Exception as pe:print(f"Panes error: {pe}")
            self._status(f"完成 — {r.total_cell_weight_g:.1f}g | {r.gravimetric_energy_Wh_kg:.0f}Wh/kg | Dcore={r.core_diameter_mm:.2f}mm")
        except Exception as e:
            import traceback;traceback.print_exc()
            messagebox.showerror("计算错误",str(e))

    def _show_summary(self,inp,r):
        tree=self.rpt_tree;tree.delete(*tree.get_children())
        pos=[];neg=[];sep=[];elec=[];core=[]
        R=pos;R.append(("目标容量",f"{inp.target_capacity_mAh:.0f}","mAh"))
        R.append(("标称电压",f"{r.nominal_voltage_V:.2f}","V"))
        R.append(("涂层真密度",f"{r.pos_coating_true_density_g_cm3:.4f}","g/cm³"))
        R.append(("压实密度",f"{r.pos_coating_pressed_density_g_cm3:.2f}","g/cm³"))
        R.append(("孔隙率",f"{r.pos_coating_porosity*100:.1f}","%"))
        R.append(("面容量",f"{r.pos_areal_capacity_mAh_cm2:.4f}","mAh/cm²"))
        R.append(("涂层厚度",f"{r.pos_coating_thickness_actual_um:.2f}","µm"))
        R.append(("涂布长度",f"{r.pos_coating_length_mm:.0f}","mm"))
        R.append(("涂布面积",f"{r.pos_coating_area_cm2:.1f}","cm²"))
        R.append(("涂层质量",f"{r.pos_coating_weight_g:.2f}","g"))
        R.append(("  活物质/SP/PVdF",f"{r.pos_active_weight_g:.2f}/{r.pos_SP_weight_g:.3f}/{r.pos_binder_weight_g:.3f}","g"))
        R.append(("Al箔质量",f"{r.pos_foil_weight_g:.2f}","g"))
        R.append(("正极总质量",f"{r.pos_total_weight_g:.2f}","g"))
        R=neg;R.append(("涂层真密度",f"{r.neg_coating_true_density_g_cm3:.4f}","g/cm³"))
        R.append(("压实密度",f"{r.neg_coating_pressed_density_g_cm3:.2f}","g/cm³"))
        R.append(("孔隙率",f"{r.neg_coating_porosity*100:.1f}","%"))
        R.append(("面容量",f"{r.neg_areal_capacity_mAh_cm2:.4f}","mAh/cm²"))
        R.append(("涂层厚度",f"{r.neg_coating_thickness_actual_um:.2f}","µm"))
        R.append(("涂布长度",f"{r.neg_coating_length_mm:.0f}","mm"))
        R.append(("涂布面积",f"{r.neg_coating_area_cm2:.1f}","cm²"))
        R.append(("涂层质量",f"{r.neg_coating_weight_g:.2f}","g"))
        R.append(("  活物质/CMC+SBR",f"{r.neg_active_weight_g:.2f}/{r.neg_CMC_weight_g+r.neg_binder_weight_g:.3f}","g"))
        R.append(("Cu箔质量",f"{r.neg_foil_weight_g:.2f}","g"))
        R.append(("负极总质量",f"{r.neg_total_weight_g:.2f}","g"))
        R.append(("N/P比(容量)",f"{r.actual_np_ratio:.4f}",f"{'✅安全' if r.is_np_safe else '⛔析锂风险'}"))
        R.append(("N/P比(面密度)",f"{r.areal_np_ratio:.4f}",""))
        R=sep;R.append(("隔膜宽/全长",f"{r.separator_width_mm:.0f}/{r.separator_total_length_mm:.0f}","mm"))
        R.append(("隔膜体积/质量",f"{r.separator_volume_cm3:.3f}/{r.separator_weight_g:.2f}","cm³/g"))
        R.append(("单元总厚/层数",f"{r.unit_thickness_um:.1f}/{r.total_winding_layers}","µm/层"))
        R.append(("螺旋线长",f"{r.total_spiral_length_mm:.0f}","mm"))
        R.append(("卷芯D/D100%",f"{r.core_diameter_mm:.2f}/{r.core_diameter_soc100_mm:.2f}","mm"))
        R.append(("直径裕量",f"{r.diameter_margin_mm:.3f}","mm"))
        R=elec;R.append(("总孔隙/注液量",f"{r.total_pore_volume_cm3:.3f}/{r.electrolyte_volume_cm3:.2f}","cm³"))
        R.append(("电解液/壳体质量",f"{r.electrolyte_weight_g:.2f}/{r.case_weight_g:.2f}","g"))
        R.append(("绝缘片/极耳/顶盖",f"{r.insulator_pos_weight_g+r.insulator_neg_weight_g+r.tab_pos_weight_g+r.tab_neg_weight_g+r.cap_weight_g:.2f}","g"))
        R=core;R.append(("电芯总质量",f"{r.total_cell_weight_g:.2f}","g"))
        R.append(("电芯能量",f"{r.cell_energy_Wh:.2f}","Wh"))
        R.append(("质量能量密度",f"{r.gravimetric_energy_Wh_kg:.0f}","Wh/kg"))
        R.append(("体积能量密度",f"{r.volumetric_energy_Wh_L:.0f}","Wh/L"))
        R.append(("入壳结果",f"{'✅ 成功' if r.is_core_fit else '⛔ 失败'}",
                   f"{'裕量'+f'{r.diameter_margin_mm:.2f}mm' if r.is_core_fit else '卷芯超限!'}"))
        TBL=[("",pos,"正极"),("",neg,"负极"),("",sep,"隔膜&卷绕"),("",elec,"电解液&辅件"),("",core,"汇总")]
        for _,rows,title in TBL:
            tree.insert("","end",values=(f"▸ {title}","",""),tags=("section",))
            for row in rows:tree.insert("","end",values=row)
        tree.tag_configure("section",font=("Microsoft YaHei UI",10,"bold"),foreground=IOS["accent"])

    def _show_panes(self,r):
        # 正极
        self.panes["pos"].delete(1.0,tk.END)
        self.panes["pos"].insert(tk.END,
            f"涂层真密度: {r.pos_coating_true_density_g_cm3:.4f} g/cm³\n"
            f"压实密度: {r.pos_coating_pressed_density_g_cm3:.2f} g/cm³\n"
            f"涂层孔隙率: {r.pos_coating_porosity*100:.2f}%\n"
            f"面容量: {r.pos_areal_capacity_mAh_cm2:.4f} mAh/cm²\n"
            f"辊压目标厚: {r.pos_coating_thickness_target_um:.2f} µm\n"
            f"回弹后实厚: {r.pos_coating_thickness_actual_um:.2f} µm\n"
            f"Al箔厚(延展后): {r.pos_total_thickness_um - 2*r.pos_coating_thickness_actual_um:.1f} µm\n"
            f"极片总厚度: {r.pos_total_thickness_um:.2f} µm\n"
            f"涂布面积: {r.pos_coating_area_cm2:.2f} cm²\n"
            f"涂布长度: {r.pos_coating_length_mm:.1f} mm\n"
            f"涂层质量: {r.pos_coating_weight_g:.3f} g\n"
            f"  活物质: {r.pos_active_weight_g:.3f} g\n"
            f"  SP: {r.pos_SP_weight_g:.4f} g\n"
            f"  PVdF: {r.pos_binder_weight_g:.4f} g\n"
            f"Al箔质量: {r.pos_foil_weight_g:.3f} g\n"
            f"极片总质量: {r.pos_total_weight_g:.3f} g\n"
            f"涂层体积: {r.pos_coating_volume_cm3:.4f} cm³\n"
            f"孔隙体积: {r.pos_pore_volume_cm3:.4f} cm³\n"
            f"正极容量: {r.pos_capacity_mAh:.1f} mAh")
        # 负极
        self.panes["neg"].delete(1.0,tk.END)
        self.panes["neg"].insert(tk.END,
            f"涂层真密度: {r.neg_coating_true_density_g_cm3:.4f} g/cm³\n"
            f"压实密度: {r.neg_coating_pressed_density_g_cm3:.2f} g/cm³\n"
            f"涂层孔隙率: {r.neg_coating_porosity*100:.2f}%\n"
            f"面容量: {r.neg_areal_capacity_mAh_cm2:.4f} mAh/cm²\n"
            f"辊压目标厚: {r.neg_coating_thickness_target_um:.2f} µm\n"
            f"回弹后实厚: {r.neg_coating_thickness_actual_um:.2f} µm\n"
            f"Cu箔厚(延展后): {r.neg_total_thickness_um - 2*r.neg_coating_thickness_actual_um:.1f} µm\n"
            f"极片总厚度: {r.neg_total_thickness_um:.2f} µm\n"
            f"极片宽度: {r.neg_width_mm:.1f} mm\n"
            f"涂布面积: {r.neg_coating_area_cm2:.2f} cm²\n"
            f"涂布长度: {r.neg_coating_length_mm:.1f} mm\n"
            f"涂层质量: {r.neg_coating_weight_g:.3f} g\n"
            f"  活物质: {r.neg_active_weight_g:.3f} g\n"
            f"  SP: {r.neg_SP_weight_g:.4f} g\n"
            f"  CMC: {r.neg_CMC_weight_g:.4f} g\n"
            f"  SBR: {r.neg_binder_weight_g:.4f} g\n"
            f"Cu箔质量: {r.neg_foil_weight_g:.3f} g\n"
            f"极片总质量: {r.neg_total_weight_g:.3f} g\n"
            f"涂层体积: {r.neg_coating_volume_cm3:.4f} cm³\n"
            f"孔隙体积: {r.neg_pore_volume_cm3:.4f} cm³\n"
            f"负极容量: {r.neg_capacity_mAh:.1f} mAh\n"
            f"N/P(容量): {r.actual_np_ratio:.4f}  N/P(面密度): {r.areal_np_ratio:.4f}")
        # 卷绕&隔膜
        self.panes["winding"].delete(1.0,tk.END)
        self.panes["winding"].insert(tk.END,
            f"单元总厚: {r.unit_thickness_um:.2f} µm = {r.unit_thickness_mm:.5f} mm\n"
            f"螺线参数a: {r.spiral_parameter_a_um:.4f} µm (T_unit/2π)\n"
            f"总卷绕层数: {r.total_winding_layers}\n"
            f"正极圈数: {r.positive_winding_turns}  负极圈数: {r.negative_winding_turns}  隔膜圈数: {r.separator_winding_turns}\n"
            f"理论螺旋线长: {r.total_spiral_length_mm:.2f} mm\n"
            f"卷芯直径: {r.core_diameter_mm:.3f} mm\n"
            f"SOC100%卷芯直径: {r.core_diameter_soc100_mm:.3f} mm\n"
            f"直径裕量: {r.diameter_margin_mm:.3f} mm\n"
            f"卷芯体积: {r.core_volume_cm3:.3f} cm³\n"
            f"隔膜宽度: {r.separator_width_mm:.1f} mm\n"
            f"隔膜全长: {r.separator_total_length_mm:.0f} mm\n"
            f"隔膜体积: {r.separator_volume_cm3:.4f} cm³\n"
            f"隔膜孔隙体积: {r.separator_pore_volume_cm3:.4f} cm³\n"
            f"隔膜质量: {r.separator_weight_g:.3f} g")
        # 汇总
        self.panes["summary"].delete(1.0,tk.END)
        self.panes["summary"].insert(tk.END,
            f"总孔隙体积: {r.total_pore_volume_cm3:.4f} cm³\n"
            f"电解液注入量: {r.electrolyte_volume_cm3:.2f} cm³ (×{self.iv['elyte_f'].get()})\n"
            f"电解液质量: {r.electrolyte_weight_g:.2f} g\n"
            f"壳体质量: {r.case_weight_g:.2f} g\n"
            f"绝缘片(正): {r.insulator_pos_weight_g:.3f} g  绝缘片(负): {r.insulator_neg_weight_g:.3f} g\n"
            f"极耳(正-Al): {r.tab_pos_weight_g:.3f} g  极耳(负-Cu): {r.tab_neg_weight_g:.3f} g\n"
            f"顶盖: {r.cap_weight_g:.3f} g\n"
            f"──────────────────────────\n"
            f"卷芯质量: {r.core_weight_g:.2f} g\n"
            f"电芯总质量: {r.total_cell_weight_g:.2f} g\n"
            f"电芯体积: {r.cell_volume_cm3:.2f} cm³\n"
            f"电芯能量: {r.cell_energy_Wh:.2f} Wh\n"
            f"质量能量密度: {r.gravimetric_energy_Wh_kg:.0f} Wh/kg\n"
            f"体积能量密度: {r.volumetric_energy_Wh_L:.0f} Wh/L\n"
            f"入壳: {'✅ 成功' if r.is_core_fit else '⛔ 失败 (SOC100%膨胀超限)'}")
        # 配方详情
        self.panes["pos_detail"].delete(1.0,tk.END)
        self.panes["pos_detail"].insert(tk.END,
            f"涂层真密度计算:\n"
            f"  1/ρ = {self.iv['pos_act_f'].get()}/{self.iv['pos_true_d'].get()} + {self.iv['pos_sp_f'].get()}/2.05 + {self.iv['pos_bind_f'].get()}/1.77\n"
            f"  1/ρ = {r.pos_active_weight_g/r.pos_coating_weight_g/4.65 if r.pos_coating_weight_g>0 else 0:.4f} + ...\n"
            f"  ρ_coat = {r.pos_coating_true_density_g_cm3:.4f} g/cm³\n"
            f"压实密度: {r.pos_coating_pressed_density_g_cm3:.2f} g/cm³\n"
            f"孔隙率: ε = 1 - {r.pos_coating_pressed_density_g_cm3:.2f}/{r.pos_coating_true_density_g_cm3:.4f} = {r.pos_coating_porosity*100:.2f}%\n"
            f"\n涂层厚度:\n"
            f"  目标 t = AD×10/PD = {self.iv['pos_ad'].get()}×10/{self.iv['pos_pd'].get()} = {r.pos_coating_thickness_target_um:.2f} µm\n"
            f"  实际 t = {r.pos_coating_thickness_target_um:.2f}×(1+{self.iv['sback'].get()}) = {r.pos_coating_thickness_actual_um:.2f} µm\n"
            f"\n涂层体积: {r.pos_coating_weight_g:.3f} / {r.pos_coating_true_density_g_cm3:.4f} = {r.pos_coating_volume_cm3:.4f} cm³\n"
            f"孔隙体积: {r.pos_coating_volume_cm3:.4f} × {r.pos_coating_porosity*100:.2f}% = {r.pos_pore_volume_cm3:.4f} cm³")
        self.panes["neg_detail"].delete(1.0,tk.END)
        self.panes["neg_detail"].insert(tk.END,
            f"涂层真密度计算:\n"
            f"  1/ρ = {self.iv['neg_act_f'].get()}/2.2 + {self.iv['neg_sp_f'].get()}/2.05 + {self.iv['neg_cmc_f'].get()}/1.28 + {self.iv['neg_bind_f'].get()}/1.0\n"
            f"  ρ_coat = {r.neg_coating_true_density_g_cm3:.4f} g/cm³\n"
            f"压实密度: {r.neg_coating_pressed_density_g_cm3:.2f} g/cm³\n"
            f"孔隙率: ε = 1 - {r.neg_coating_pressed_density_g_cm3:.2f}/{r.neg_coating_true_density_g_cm3:.4f} = {r.neg_coating_porosity*100:.2f}%\n"
            f"\n涂层厚度:\n"
            f"  目标 t = AD×10/PD = {self.iv['neg_ad'].get()}×10/{self.iv['neg_pd'].get()} = {r.neg_coating_thickness_target_um:.2f} µm\n"
            f"  实际 t = {r.neg_coating_thickness_target_um:.2f}×(1+{self.iv['sback'].get()}) = {r.neg_coating_thickness_actual_um:.2f} µm\n"
            f"\n膨胀后负极厚(SOC100%): {r.neg_total_thickness_um*(1+float(self.iv['neg_exp'].get())):.2f} µm\n"
            f"涂层体积: {r.neg_coating_weight_g:.3f}/{r.neg_coating_true_density_g_cm3:.4f} = {r.neg_coating_volume_cm3:.4f} cm³\n"
            f"孔隙体积: {r.neg_coating_volume_cm3:.4f}×{r.neg_coating_porosity*100:.2f}% = {r.neg_pore_volume_cm3:.4f} cm³")

    def _export(self):
        if self.result is None:messagebox.showwarning("无数据","请先计算");return
        fp=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],initialfile=f"{self.cell_type.get()}_design.xlsx")
        if not fp:return
        try:
            import openpyxl;from openpyxl.styles import Font as XF,PatternFill,Alignment,Border,Side
            wb=openpyxl.Workbook();r=self.result;hf=XF(name="Microsoft YaHei",size=10,bold=True,color="FFFFFF")
            hfl=PatternFill(start_color="007AFF",end_color="007AFF",fill_type="solid")
            bd=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
            data=[("参数","数值","单位"),
                  ("正极涂布长度",f"{r.pos_coating_length_mm:.1f}","mm"),
                  ("正极涂布面积",f"{r.pos_coating_area_cm2:.1f}","cm²"),
                  ("正极涂层厚度",f"{r.pos_coating_thickness_actual_um:.1f}","µm"),
                  ("正极总厚度",f"{r.pos_total_thickness_um:.1f}","µm"),
                  ("正极孔隙率",f"{r.pos_coating_porosity*100:.1f}","%"),
                  ("正极活物质质量",f"{r.pos_active_weight_g:.2f}","g"),
                  ("正极总质量",f"{r.pos_total_weight_g:.2f}","g"),
                  ("负极涂布长度",f"{r.neg_coating_length_mm:.1f}","mm"),
                  ("负极涂层厚度",f"{r.neg_coating_thickness_actual_um:.1f}","µm"),
                  ("负极孔隙率",f"{r.neg_coating_porosity*100:.1f}","%"),
                  ("N/P(容量)",f"{r.actual_np_ratio:.4f}",""),
                  ("N/P(面密度)",f"{r.areal_np_ratio:.4f}",""),
                  ("单元总厚",f"{r.unit_thickness_um:.1f}","µm"),
                  ("卷绕层数",r.total_winding_layers,"层"),
                  ("卷芯直径",f"{r.core_diameter_mm:.2f}","mm"),
                  ("SOC100%直径",f"{r.core_diameter_soc100_mm:.2f}","mm"),
                  ("螺旋线长",f"{r.total_spiral_length_mm:.0f}","mm"),
                  ("总孔隙体积",f"{r.total_pore_volume_cm3:.3f}","cm³"),
                  ("注液量",f"{r.electrolyte_volume_cm3:.2f}","cm³"),
                  ("电解液质量",f"{r.electrolyte_weight_g:.2f}","g"),
                  ("电芯总质量",f"{r.total_cell_weight_g:.2f}","g"),
                  ("能量密度(质量)",f"{r.gravimetric_energy_Wh_kg:.0f}","Wh/kg"),
                  ("能量密度(体积)",f"{r.volumetric_energy_Wh_L:.0f}","Wh/L"),
            ]
            ws=wb.active;ws.title="设计结果"
            ws.merge_cells('A1:C1');ws['A1']=f"{self.cell_type.get()} 全参数设计报告";ws['A1'].font=XF(name="Microsoft YaHei",size=14,bold=True,color="007AFF")
            for i,row in enumerate(data):
                for j,v in enumerate(row):
                    c=ws.cell(row=i+3,column=j+1,value=v)
                    c.font=XF(name="Microsoft YaHei",size=10,bold=True)if i==0 else XF(name="Microsoft YaHei",size=10)
                    if i==0:c.font=hf;c.fill=hfl
                    c.border=bd;c.alignment=Alignment(horizontal='center')
            for col in range(1,4):ws.column_dimensions[chr(64+col)].width=22
            wb.save(fp);self._status(f"已导出: {fp}");messagebox.showinfo("完成",f"Excel已保存:\n{fp}")
        except Exception as e:messagebox.showerror("导出错误",str(e))

    def _status(self,msg):self.status.configure(text=f"  {msg}");self.root.update_idletasks()


def main():root=tk.Tk();App(root);root.mainloop()
if __name__=="__main__":main()
